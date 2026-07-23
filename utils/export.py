from __future__ import annotations

import io
from copy import copy

import pandas as pd
from openpyxl import load_workbook


BASE_COLUMNS = ["ID", "Nome", "Cognome", "Struttura"]


def _norm(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _python_value(value: object):
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


def _find_data_sheet(workbook):
    required = {_norm(c) for c in BASE_COLUMNS}
    for worksheet in workbook.worksheets:
        for row_idx in range(1, min(worksheet.max_row, 20) + 1):
            headers = {
                _norm(worksheet.cell(row=row_idx, column=col_idx).value)
                for col_idx in range(1, worksheet.max_column + 1)
            }
            if required.issubset(headers):
                return worksheet, row_idx
    return workbook.active, 1


def update_excel_bytes(source_bytes: bytes, df: pd.DataFrame) -> bytes:
    """Aggiorna il foglio dati del file caricato preservando gli altri fogli e la formattazione."""
    workbook = load_workbook(io.BytesIO(source_bytes))
    worksheet, header_row = _find_data_sheet(workbook)

    existing_headers: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=header_row, column=col_idx).value
        if value is not None and str(value).strip():
            existing_headers[str(value).strip()] = col_idx

    # Aggiunge in coda soltanto le colonne mancanti, copiando lo stile dell'ultima intestazione.
    last_header_col = max(existing_headers.values(), default=0)
    for column in df.columns:
        if column in existing_headers:
            continue
        last_header_col += 1
        target = worksheet.cell(row=header_row, column=last_header_col, value=column)
        if last_header_col > 1:
            source = worksheet.cell(row=header_row, column=last_header_col - 1)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.fill:
                target.fill = copy(source.fill)
            if source.font:
                target.font = copy(source.font)
            if source.border:
                target.border = copy(source.border)
        existing_headers[column] = last_header_col

    for row_offset, (_, row) in enumerate(df.iterrows(), start=1):
        excel_row = header_row + row_offset
        for column in df.columns:
            worksheet.cell(
                row=excel_row,
                column=existing_headers[column],
                value=_python_value(row[column]),
            )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
