from __future__ import annotations

from io import BytesIO
import re

import pandas as pd
from openpyxl import load_workbook

from .recode import normalize_level


BASE_COLUMNS = ["ID", "Nome", "Cognome", "Struttura"]


def _norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def _read_bytes(file) -> bytes:
    if isinstance(file, (bytes, bytearray)):
        return bytes(file)
    if hasattr(file, "getvalue"):
        return file.getvalue()
    if hasattr(file, "read"):
        position = file.tell() if hasattr(file, "tell") else None
        raw = file.read()
        if position is not None and hasattr(file, "seek"):
            file.seek(position)
        return raw
    raise TypeError("Formato file non supportato")


def _find_data_sheet(raw: bytes) -> tuple[str, int]:
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=False)
    required = {_norm(c) for c in BASE_COLUMNS}
    try:
        for worksheet in workbook.worksheets:
            max_scan = min(worksheet.max_row or 1, 25)
            for row_idx in range(1, max_scan + 1):
                headers = {
                    _norm(worksheet.cell(row=row_idx, column=col_idx).value)
                    for col_idx in range(1, (worksheet.max_column or 1) + 1)
                }
                if required.issubset(headers):
                    return worksheet.title, row_idx
    finally:
        workbook.close()
    raise ValueError("Nessun foglio contiene le intestazioni ID, Nome, Cognome e Struttura")


def read_department_excel(file) -> pd.DataFrame:
    """Legge il foglio anagrafico anche quando non è il primo o l'intestazione non è in riga 1."""
    raw = _read_bytes(file)
    sheet_name, header_row = _find_data_sheet(raw)
    df = pd.read_excel(BytesIO(raw), sheet_name=sheet_name, header=header_row - 1, dtype=object)
    df = df.rename(columns={c: str(c).strip() for c in df.columns})
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def normalize_competence_columns(df: pd.DataFrame, codes: list[str]) -> pd.DataFrame:
    out = df.copy()
    missing = [code for code in codes if code not in out.columns]
    if missing:
        additions = pd.DataFrame("NA", index=out.index, columns=missing)
        out = pd.concat([out, additions], axis=1)
    for code in codes:
        out[code] = out[code].apply(normalize_level)
    return out.copy()


def detect_structure_values(df: pd.DataFrame) -> list[str]:
    if "Struttura" not in df.columns:
        return []
    values = df["Struttura"].dropna().astype(str).map(str.strip)
    return sorted(v for v in values.unique() if v)
